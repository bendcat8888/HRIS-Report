# CREATED DATE: 03-17-2025 (Updated from 01-23-2023)
# USER: HR Staff for leave Computation
from colorama import Fore, Back, Style
import pyodbc
import os
import pandas as pd
import getpass
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def create_pivot_table(input_csv, output_csv=None):
    """
    Process the IHRIS leave data into a pivoted format with leave types as columns
    
    Parameters:
    input_csv (str): Path to the input CSV file
    output_csv (str, optional): Path for the output CSV file. If not provided,
                               a default name is generated based on the input file.
    """
    print(f"Creating pivot table from: {input_csv}")
    
    # Read the CSV file
    try:
        df = pd.read_csv(input_csv)
    except Exception as e:
        print(f"Error reading CSV file: {e}")
        return
    
    # Fill NaN values with 0 for numeric columns
    df['daywithpay'] = df['daywithpay'].fillna(0)
    df['daywopay'] = df['daywopay'].fillna(0)
    
    # Group by employee and pivot the data
    employee_data = []
    
    # Get unique employee data with department name
    employees = df[['group_code', 'emp_code', 'emp_name', 'dept_name']].drop_duplicates().values.tolist()
    
    # Process each employee
    for emp in employees:
        group_code, emp_code, emp_name, dept_name = emp
        employee_rows = df[(df['emp_code'] == emp_code)]
        
        # Initialize employee row with all leave types set to 0
        emp_data = {
            'group_code': group_code,
            'emp_code': emp_code,
            'emp_name': emp_name,
            'dept_name': dept_name,
            'AW_daywithpay': 0,
            'AW_daywopay': 0,
            'SL_daywithpay': 0,
            'SL_daywopay': 0,
            'VL_daywithpay': 0,
            'VL_daywopay': 0
        }
        
        # Fill in leave data
        for _, row in employee_rows.iterrows():
            leave_code = row.get('leavecode')
            if leave_code in ['AW', 'SL', 'VL']:  # Only process the main leave types we need
                emp_data[f"{leave_code}_daywithpay"] = row['daywithpay'] if row['daywithpay'] > 0 else 0
                emp_data[f"{leave_code}_daywopay"] = row['daywopay'] if row['daywopay'] > 0 else 0
        
        # Calculate total
        emp_data['Total'] = (
            emp_data['AW_daywithpay'] + emp_data['AW_daywopay'] +
            emp_data['SL_daywithpay'] + emp_data['SL_daywopay'] +
            emp_data['VL_daywithpay'] + emp_data['VL_daywopay']
        )
        
        employee_data.append(emp_data)
    
    # Create DataFrame from processed data
    pivot_df = pd.DataFrame(employee_data)
    
    # Sort by group_code and emp_code
    pivot_df = pivot_df.sort_values(by=['group_code', 'emp_code'])
    
    # Define column order
    column_order = [
        'group_code', 'emp_code', 'emp_name', 'dept_name',
        'AW_daywithpay', 'AW_daywopay',
        'SL_daywithpay', 'SL_daywopay',
        'VL_daywithpay', 'VL_daywopay',
        'Total'
    ]
    
    # Reorder columns
    pivot_df = pivot_df[column_order]
    
    # Generate output file name if not provided
    if output_csv is None:
        base_name = os.path.splitext(input_csv)[0]
        output_csv = f"{base_name}_pivot.csv"
    
    # Save to CSV
    pivot_df.to_csv(output_csv, index=False)
    print(f"Pivot table saved to: {output_csv}")
    return pivot_df

def create_detailed_report(results_df, output_file=None, start_date=None, end_date=None):
    """
    Create a detailed Excel report with complex header structure matching the template.
    Row 4 (first data row): Field names for columns
    Row 8: Additional conditions to get values for columns
    Other columns remain blank.
    
    Parameters:
    results_df (DataFrame): DataFrame containing the SQL query results
    output_file (str, optional): Path for the output Excel file
    start_date (str, optional): Start date for filename
    end_date (str, optional): End date for filename
    """
    print("Creating detailed report...")
    
    # Generate output file name if not provided
    if output_file is None:
        if start_date and end_date:
            output_file = f"IHRIS_Detailed_Report_{start_date}_to_{end_date}.xlsx"
        else:
            output_file = "IHRIS_Detailed_Report.xlsx"
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "IHRIS Report"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Row 1: Top-level categories
    # A=Company Name, B=CODE (emp_code), C=Employee Name, D=Date Hired, E=Department, F=Division, G-H=TARDINESS
    # W/O LEAVE FORM spans I-K (columns 9-11)
    ws.merge_cells('I1:K1')
    ws['I1'] = 'W/O LEAVE FORM'
    ws['I1'].fill = header_fill
    ws['I1'].font = header_font
    ws['I1'].alignment = center_align
    ws['I1'].border = thin_border
    
    # W/ LEAVE FORM spans L-O (columns 12-15)
    ws.merge_cells('L1:O1')
    ws['L1'] = 'W/ LEAVE FORM'
    ws['L1'].fill = header_fill
    ws['L1'].font = header_font
    ws['L1'].alignment = center_align
    ws['L1'].border = thin_border
    
    # SPECIAL LEAVE spans P-T (columns 16-20)
    ws.merge_cells('P1:T1')
    ws['P1'] = 'SPECIAL LEAVE'
    ws['P1'].fill = header_fill
    ws['P1'].font = header_font
    ws['P1'].alignment = center_align
    ws['P1'].border = thin_border
    
    # Apply borders to merged cells in row 1 (A through V)
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V']:
        cell = ws[f'{col}1']
        if cell.value is None:
            cell.border = thin_border
    
    # Row 2: Main column headers
    # A=Company Name (group_code), B=CODE (emp_code), C=EMPLOYEE NAME, D=DATE HIRED, E=DEPARTMENT, F=Division (sect_name)
    # G-H: TARDINESS (merged)
    # I-K: W/O LEAVE FORM (I=ABSENT, J=HALFDAY-TARDINESS, K=SUSPENSION)
    # L-O: W/ LEAVE FORM (L-M=SICK LEAVE merged, N-O=VACATION LEAVE merged)
    # P-T: SPECIAL LEAVE (P=MATERNITY LEAVE, Q=SOLO PARENT LEAVE, R=PATERNITY LEAVE, S=BEREAVEMENT LEAVE, T=MAGNA CARTA FOR WOMEN)
    # U=OFFSET, V=TOTAL
    
    main_headers = {
        'A2': 'COMPANY NAME',
        'B2': 'CODE',
        'C2': 'EMPLOYEE NAME',
        'D2': 'DATE HIRED',
        'E2': 'DEPARTMENT',
        'F2': 'DIVISION',
        'G2': 'TARDINESS',  # Will merge G2:H2
        'I2': 'ABSENT',
        'J2': 'HALFDAY-TARDINESS',
        'K2': 'SUSPENSION',
        'L2': 'SICK LEAVE',  # Will merge L2:M2
        'N2': 'VACATION LEAVE',  # Will merge N2:O2
        'P2': 'MATERNITY LEAVE',
        'Q2': 'SOLO PARENT LEAVE',
        'R2': 'PATERNITY LEAVE',
        'S2': 'BEREAVEMENT LEAVE',
        'T2': 'MAGNA CARTA FOR WOMEN',
        'U2': 'OFFSET',
        'V2': 'TOTAL'
    }
    
    for cell, text in main_headers.items():
        ws[cell] = text
        ws[cell].fill = subheader_fill
        ws[cell].font = subheader_font
        ws[cell].alignment = center_align
        ws[cell].border = thin_border
    
    # Merge cells for multi-column headers
    ws.merge_cells('G2:H2')  # TARDINESS
    ws.merge_cells('L2:M2')  # SICK LEAVE
    ws.merge_cells('N2:O2')  # VACATION LEAVE
    
    # Apply borders to all cells in row 2 (columns 1-22)
    for col in range(1, 23):
        cell = ws.cell(row=2, column=col)
        if cell.value is None:
            cell.border = thin_border
    
    # Row 3: Sub-headers
    subheaders = {
        'A3': '',
        'B3': '',
        'C3': '',
        'D3': '',
        'E3': '',
        'F3': '',
        'G3': 'MINUTES',
        'H3': 'FREQUENCY',
        'I3': '',
        'J3': '',
        'K3': '',
        'L3': 'W/ PAY',
        'M3': 'W/O PAY',
        'N3': 'W/ PAY',
        'O3': 'W/O PAY',
        'P3': '',  # MATERNITY LEAVE
        'Q3': '',
        'R3': '',
        'S3': '',
        'T3': '',
        'U3': '',
        'V3': ''
    }
    
    for cell, text in subheaders.items():
        ws[cell] = text
        ws[cell].font = normal_font
        ws[cell].alignment = center_align
        ws[cell].border = thin_border
        # Apply subheader fill to subheader cells
        if text:  # Only fill cells with text
            ws[cell].fill = subheader_fill
    
    # Set column widths
    column_widths = {
        'A': 18,  # Company Name (group_code)
        'B': 12,  # CODE (emp_code)
        'C': 25,  # EMPLOYEE NAME
        'D': 12,  # DATE HIRED
        'E': 20,  # DEPARTMENT
        'F': 18,  # Division (sect_name)
        'G': 10,  # TARDINESS MINUTES
        'H': 10,  # TARDINESS FREQUENCY
        'I': 12,  # ABSENT
        'J': 18,  # HALFDAY-TARDINESS
        'K': 20,  # SUSPENSION
        'L': 12,  # SICK LEAVE W/ PAY
        'M': 12,  # SICK LEAVE W/O PAY
        'N': 12,  # VACATION LEAVE W/ PAY
        'O': 12,  # VACATION LEAVE W/O PAY
        'P': 20,  # MATERNITY LEAVE
        'Q': 30,  # SOLO PARENT LEAVE
        'R': 30,  # PATERNITY LEAVE
        'S': 30,  # BEREAVEMENT LEAVE
        'T': 20,  # MAGNA CARTA FOR WOMEN
        'U': 12,  # OFFSET
        'V': 12   # TOTAL
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Set row heights
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    
    # Process data and populate rows starting from row 4
    if not results_df.empty:
        # Fill NaN values
        results_df['daywithpay'] = results_df['daywithpay'].fillna(0)
        results_df['daywopay'] = results_df['daywopay'].fillna(0)
        results_df['reason'] = results_df['reason'].fillna('')
        
        # Get unique employees (include group_code and sect_name for new columns A and E)
        employees = results_df[['emp_code', 'emp_name', 'hired_date', 'dept_name', 'group_code', 'sect_name']].drop_duplicates(subset=['emp_code'])
        
        row_num = 4  # Start data rows after headers (row 3)
        
        for _, emp in employees.iterrows():
            emp_code = emp['emp_code']
            emp_data = results_df[results_df['emp_code'] == emp_code]
            # add OFFSET column
            
            # Initialize employee row data
            emp_row = {
                'group_code': emp['group_code'],
                'emp_name': emp['emp_name'],
                'hired_date': emp['hired_date'],
                'dept_name': emp['dept_name'],
                'sect_name': emp['sect_name'] if pd.notna(emp.get('sect_name')) else '',
                'AW_daywopay': 0,  # ABSENT - W/O PAY
                'SL_daywithpay': 0,  # SICK LEAVE W/ PAY
                'SL_daywopay': 0,  # SICK LEAVE W/O PAY
                'VL_daywithpay': 0,  # VACATION LEAVE W/ PAY
                'VL_daywopay': 0,  # VACATION LEAVE W/O PAY
                'SU_total': 0,  # SUSPENSION (leave_code = 'SU')
                'MN_daywopay': 0,  # MATERNITY LEAVE W/O PAY (leave_code = 'MN')
                'MN_sss': 0,  # MATERNITY LEAVE (C/O SSS)
                'SoloParent': 0,  # SOLO PARENT LEAVE
                'Paternity': 0,  # PATERNITY LEAVE
                'Bereavement': 0,  # BEREAVEMENT LEAVE
                'OFFSET': 0  # OFFSET
            }
            
            # Process leave records for this employee
            for _, record in emp_data.iterrows():
                leavecode = record.get('leavecode', '')
                reason = str(record.get('reason', '')).upper()
                daywithpay = float(record.get('daywithpay', 0) or 0)
                daywopay = float(record.get('daywopay', 0) or 0)
                
                # ABSENT - W/O PAY (leavecode = 'AW' and daywopay > 0)
                if leavecode == 'AW' and daywopay > 0:
                    emp_row['AW_daywopay'] += daywopay
                
                # SICK LEAVE
                if leavecode == 'SL':
                    if daywithpay > 0:
                        emp_row['SL_daywithpay'] += daywithpay
                    if daywopay > 0:
                        emp_row['SL_daywopay'] += daywopay
                
                # VACATION LEAVE
                if leavecode == 'VL':
                    # Check for special leave types first
                    if 'SOLO PARENT' in reason or 'SP:' in reason:
                        emp_row['SoloParent'] += (daywithpay + daywopay)
                    elif 'PL:' in reason:
                        emp_row['Paternity'] += (daywithpay + daywopay)
                    elif 'BL:' in reason:
                        emp_row['Bereavement'] += (daywithpay + daywopay)
                    # elif 'OFFSET:' in reason:
                    #     emp_row['OFFSET'] += (daywithpay + daywopay)
                    # Regular vacation leave
                    else:
                        # Regular vacation leave
                        if daywithpay > 0:
                            emp_row['VL_daywithpay'] += daywithpay
                        if daywopay > 0:
                            emp_row['VL_daywopay'] += daywopay
                # OFFSET (leavecode = 'OB')
                if leavecode == 'OB' and 'OFFSET:' in reason:
                    emp_row['OFFSET'] += (daywithpay + daywopay)
                
                # SUSPENSION (leavecode = 'SU')
                if leavecode == 'SU':
                    emp_row['SU_total'] += (daywithpay + daywopay)
                
                # MATERNITY LEAVE (leavecode = 'MN')
                if leavecode == 'MN':
                    if daywopay > 0:
                        emp_row['MN_daywopay'] += daywopay
                    # Check for C/O SSS - this might need adjustment based on actual data structure
                    # Assuming C/O SSS is when there's a specific indicator in reason or other field
                    # For now, we'll use daywithpay as indicator for C/O SSS
                    if daywithpay > 0:
                        emp_row['MN_sss'] += daywithpay
            
            # Write employee row to Excel
            # Column A: Company Name (group_code)
            ws.cell(row=row_num, column=1).value = emp_row['group_code'] if pd.notna(emp_row.get('group_code')) else ''
            ws.cell(row=row_num, column=1).border = thin_border
            ws.cell(row=row_num, column=1).alignment = left_align
            
            # Column B: CODE (emp_code)
            ws.cell(row=row_num, column=2).value = emp_code if pd.notna(emp_code) else ''
            ws.cell(row=row_num, column=2).border = thin_border
            ws.cell(row=row_num, column=2).alignment = left_align
            
            # Column C: EMPLOYEE NAME
            ws.cell(row=row_num, column=3).value = emp_row['emp_name']
            ws.cell(row=row_num, column=3).border = thin_border
            ws.cell(row=row_num, column=3).alignment = left_align
            
            # Column D: DATE HIRED
            if pd.notna(emp_row['hired_date']):
                hired_date = emp_row['hired_date']
                # Format date if it's a datetime object
                if isinstance(hired_date, pd.Timestamp):
                    ws.cell(row=row_num, column=4).value = hired_date.strftime('%m/%d/%Y')
                elif isinstance(hired_date, str) and hired_date.strip():
                    ws.cell(row=row_num, column=4).value = hired_date
                else:
                    ws.cell(row=row_num, column=4).value = str(hired_date)
            ws.cell(row=row_num, column=4).border = thin_border
            ws.cell(row=row_num, column=4).alignment = left_align
            
            # Column E: DEPARTMENT
            ws.cell(row=row_num, column=5).value = emp_row['dept_name']
            ws.cell(row=row_num, column=5).border = thin_border
            ws.cell(row=row_num, column=5).alignment = left_align
            
            # Column F: Division (sect_name)
            ws.cell(row=row_num, column=6).value = emp_row['sect_name']
            ws.cell(row=row_num, column=6).border = thin_border
            ws.cell(row=row_num, column=6).alignment = left_align
            
            # Column G-H: TARDINESS (MINUTES, FREQUENCY) - blank for now
            ws.cell(row=row_num, column=7).border = thin_border
            ws.cell(row=row_num, column=8).border = thin_border
            
            # Column I: ABSENT - W/O PAY
            if emp_row['AW_daywopay'] > 0:
                cell = ws.cell(row=row_num, column=9)
                cell.value = round(emp_row['AW_daywopay'], 2)
                cell.number_format = '0.00'
            ws.cell(row=row_num, column=9).border = thin_border
            ws.cell(row=row_num, column=9).alignment = center_align
            
            # Column J: HALFDAY-TARDINESS - blank for now
            ws.cell(row=row_num, column=10).border = thin_border
            
            # Column K: SUSPENSION
            if emp_row['SU_total'] > 0:
                cell = ws.cell(row=row_num, column=11)
                cell.value = round(emp_row['SU_total'], 2)
                cell.number_format = '0.00'
            ws.cell(row=row_num, column=11).border = thin_border
            ws.cell(row=row_num, column=11).alignment = center_align
            
            # Column L: SICK LEAVE W/ PAY
            if emp_row['SL_daywithpay'] > 0:
                cell = ws.cell(row=row_num, column=12)
                cell.value = round(emp_row['SL_daywithpay'], 2)
                cell.number_format = '0.00'
            ws.cell(row=row_num, column=12).border = thin_border
            ws.cell(row=row_num, column=12).alignment = center_align
            
            # Column M: SICK LEAVE W/O PAY
            if emp_row['SL_daywopay'] > 0:
                cell = ws.cell(row=row_num, column=13)
                cell.value = round(emp_row['SL_daywopay'], 2)
                cell.number_format = '0.00'
            ws.cell(row=row_num, column=13).border = thin_border
            ws.cell(row=row_num, column=13).alignment = center_align
            
            # Column N: VACATION LEAVE W/ PAY
            if emp_row['VL_daywithpay'] > 0:
                cell = ws.cell(row=row_num, column=14)
                cell.value = round(emp_row['VL_daywithpay'], 2)
                cell.number_format = '0.00'
            ws.cell(row=row_num, column=14).border = thin_border
            ws.cell(row=row_num, column=14).alignment = center_align
            
            # Column O: VACATION LEAVE W/O PAY
            if emp_row['VL_daywopay'] > 0:
                cell = ws.cell(row=row_num, column=15)
                cell.value = round(emp_row['VL_daywopay'], 2)
                cell.number_format = '0.00'
            ws.cell(row=row_num, column=15).border = thin_border
            ws.cell(row=row_num, column=15).alignment = center_align
            
            # Column P: MATERNITY LEAVE W/O PAY
            if emp_row['MN_daywopay'] > 0:
                cell = ws.cell(row=row_num, column=16)
                cell.value = round(emp_row['MN_daywopay'], 2)
                cell.number_format = '0.00'
            ws.cell(row=row_num, column=16).border = thin_border
            ws.cell(row=row_num, column=16).alignment = center_align
            
            # Column Q: SOLO PARENT LEAVE
            if emp_row['SoloParent'] > 0:
                cell = ws.cell(row=row_num, column=17)
                cell.value = round(emp_row['SoloParent'], 2)
                cell.number_format = '0.00'
            ws.cell(row=row_num, column=17).border = thin_border
            ws.cell(row=row_num, column=17).alignment = center_align
            
            # Column R: PATERNITY LEAVE
            if emp_row['Paternity'] > 0:
                cell = ws.cell(row=row_num, column=18)
                cell.value = round(emp_row['Paternity'], 2)
                cell.number_format = '0.00'
            ws.cell(row=row_num, column=18).border = thin_border
            ws.cell(row=row_num, column=18).alignment = center_align
            
            # Column S: BEREAVEMENT LEAVE
            if emp_row['Bereavement'] > 0:
                cell = ws.cell(row=row_num, column=19)
                cell.value = round(emp_row['Bereavement'], 2)
                cell.number_format = '0.00'
            ws.cell(row=row_num, column=19).border = thin_border
            ws.cell(row=row_num, column=19).alignment = center_align
            
            # Column T: MAGNA CARTA FOR WOMEN - blank
            ws.cell(row=row_num, column=20).border = thin_border
            
            # Column U: OFFSET #ADD THE CODE
            if emp_row['OFFSET'] > 0:
                cell = ws.cell(row=row_num, column=21)
                cell.value = round(emp_row['OFFSET'], 2)
                cell.number_format = '0.00'
            ws.cell(row=row_num, column=21).border = thin_border
            ws.cell(row=row_num, column=21).alignment = center_align
            # Column V: TOTAL
            total = (emp_row['AW_daywopay'] + emp_row['SL_daywithpay'] + emp_row['SL_daywopay'] +
                    emp_row['VL_daywithpay'] + emp_row['VL_daywopay'] + emp_row['SU_total'] +
                    emp_row['MN_daywopay'] + emp_row['MN_sss'] + emp_row['SoloParent'] +
                    emp_row['Paternity'] + emp_row['Bereavement'] + emp_row['OFFSET'])
            if total > 0:
                cell = ws.cell(row=row_num, column=22)
                cell.value = round(total, 2)
                cell.number_format = '0.00'
            ws.cell(row=row_num, column=22).border = thin_border
            ws.cell(row=row_num, column=22).alignment = center_align
            
            row_num += 1
    
    # Save the workbook
    wb.save(output_file)
    print(f"Detailed report with data saved to: {output_file}")
    return output_file

def doQuery(conn):
    conn.cursor()
    os.system("cls" or "clear")
    print(Back.GREEN + ">>> Link has been established <<<")
    print(Style.RESET_ALL)
    print("_________________________________________________")
    n=input("\nPlease input the STARTING date (YYYY-MM-DD) : " + Fore.BLUE + Back.WHITE + Style.BRIGHT )
    print(Style.RESET_ALL)
    m=input("Please input the ENDING date (YYYY-MM-DD) : " + Fore.BLUE + Back.WHITE + Style.BRIGHT )
    print(Style.RESET_ALL)
    
    # Add group code filtering option
    g=input("\nFilter by group codes (comma-separated, leave blank for all): " + Fore.BLUE + Back.WHITE + Style.BRIGHT )
    print(Style.RESET_ALL)
    
    # Create the base query to get all active employees
    sql = """SELECT 
        b.[group_code],
        b.[emp_code],
        b.[emp_name],
        b.[hired_date],
        a.[leavecode],
        f.[leavedesc],
        COALESCE(sum(a.[daywithpay]), 0) as [daywithpay],
        COALESCE(sum(a.[daywopay]), 0) as [daywopay],
        c.[empstatus_desc],
        d.[dept_name],
        e.[sect_name],
        a.[reason],
        NULL as [ob_reason]
    FROM 
        (SELECT [sect_code],[dept_code],[emp_code], [group_code], 
         ([last_name] + ', ' + [first_name]) as [emp_name], 
         [emp_status],[terminate_date],[hired_date]
         FROM employee
         WHERE [terminate_date] IS NULL OR [terminate_date] = '') b 
    LEFT JOIN 
        (SELECT * FROM t_leave 
         WHERE [trx_date] BETWEEN convert(date,'{0}') AND convert(date,'{1}')) a 
        ON a.[emp_code] = b.[emp_code]
    LEFT JOIN employee_status c ON b.[emp_status] = c.[empstatus_code]
    LEFT JOIN department d ON b.[dept_code] = d.[dept_code]
    LEFT JOIN section e ON b.[sect_code] = e.[sect_code]
    LEFT JOIN leave f ON a.[leavecode] = f.[leavecode]
    """.format(n, m)
    
    ob_sql = """SELECT 
        b.[group_code],
        b.[emp_code],
        b.[emp_name],
        b.[hired_date],
        'OB' as [leavecode],
        'Official Business' as [leavedesc],
        COUNT(1) as [daywithpay],
        CAST(0 as float) as [daywopay],
        c.[empstatus_desc],
        d.[dept_name],
        e.[sect_name],
        g.[reason] as [reason],
        g.[reason] as [ob_reason]
    FROM 
        (SELECT [sect_code],[dept_code],[emp_code], [group_code], 
         ([last_name] + ', ' + [first_name]) as [emp_name], 
         [emp_status],[terminate_date],[hired_date]
         FROM employee
         WHERE [terminate_date] IS NULL OR [terminate_date] = '') b 
    INNER JOIN 
        (SELECT * FROM t_ob 
         WHERE [trx_date] BETWEEN convert(date,'{0}') AND convert(date,'{1}')) g 
        ON g.[emp_code] = b.[emp_code]
    LEFT JOIN employee_status c ON b.[emp_status] = c.[empstatus_code]
    LEFT JOIN department d ON b.[dept_code] = d.[dept_code]
    LEFT JOIN section e ON b.[sect_code] = e.[sect_code]
    """.format(n, m)
    
    # Add group_code filter if provided
    if g.strip():
        group_codes = [f"'{code.strip()}'" for code in g.split(',') if code.strip()]
        if group_codes:
            group_filter = "WHERE b.[group_code] IN ({0})".format(",".join(group_codes))
            sql += group_filter
            ob_sql += group_filter
    
    # Complete the query with GROUP BY and ORDER BY
    sql += """GROUP BY 
        b.[group_code],
        b.[emp_code],
        b.[emp_name],
        b.[hired_date],
        a.[leavecode],
        c.[empstatus_desc],
        d.[dept_name],
        e.[sect_name],
        f.[leavedesc],
        a.[reason]
    ORDER BY 
        b.[group_code],
        b.[emp_name]

    """

    ob_sql += """GROUP BY 
        b.[group_code],
        b.[emp_code],
        b.[emp_name],
        b.[hired_date],
        c.[empstatus_desc],
        d.[dept_name],
        e.[sect_name],
        g.[reason]
    """

    # Create appropriate file name suffix based on filters
    filter_suffix = ""
    if g.strip():
        # Create a short representation of the group codes for the filename
        group_suffix = g.replace(' ', '').replace(',', '-')
        filter_suffix = f"_groups_{group_suffix}"
    
    results_leave = pd.read_sql_query(sql, conn)
    results_ob = pd.read_sql_query(ob_sql, conn)
    results = pd.concat([results_leave, results_ob], ignore_index=True)
    
    results['leavecode'] = results['leavecode'].fillna('NONE')
    results['leavedesc'] = results['leavedesc'].fillna('No Leave')
    
    summary_file = f"IHRIS_Summary_VL-SL_{n}_to_{m}{filter_suffix}.csv"
    pivot_file = f"IHRIS_Pivot_VL-SL_{n}_to_{m}{filter_suffix}.csv"
    detailed_file = f"IHRIS_Detailed_Report_{n}_to_{m}{filter_suffix}.xlsx"

    try:
        # Save the original query results
        results.to_csv(summary_file, index=False)
        print(f"Original summary saved to: {summary_file}")
        
        # Create the pivot table automatically
        create_pivot_table(summary_file, pivot_file)
        
        # Create the detailed report template
        create_detailed_report(results, detailed_file, n, m)

        print("\n+++++++++++++++++++++++++++")
        print("Fetched Lines [ALL]  :", len(results))
        print("Pivot table created successfully!")
        print("Detailed report template created successfully!")
        print("+++++++++++++++++++++++++++")
        
        # No longer asking if user wants to open the file, just notify them
        print(f"\nSummary file: {summary_file}")
        print(f"Pivot file: {pivot_file}")
        print(f"Detailed report file: {detailed_file}")
        
        input("\nDONE! \nPress the <ENTER> key to EXIT")

    except (IOError,NameError) as e:
        e_r = "Please close the CSV or Excel file that was shown above" if 'Errno 13' in str(e) else "Please read the error was shown above"
        txt = """

        ERROR!!!
        ::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::\n
        '{0}'

         Program was halt due to an error.
        '{1}'
        ::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::\n\n
        Press any key to Exit
        """.format(str(e),e_r)
        input(txt)


print(">>> Connecting to Server >>>")
print("This may take a while for processing the secure access..")

host = os.environ.get("IHRIS_SERVER") or input("SQL Server host (e.g., IHRIS-SERVER\\IHRISSERVER): ")
data = os.environ.get("IHRIS_DB") or input("Database name: ")
us_r = os.environ.get("IHRIS_USER") or input("Username: ")
pas_ = os.environ.get("IHRIS_PASSWORD") or getpass.getpass("Password: ")

try:
    connection_string = f"DRIVER={{SQL Server}};Server={host};PORT=1433;Database={data};uid={us_r};pwd={pas_};"
    myConnection = pyodbc.connect(connection_string)
    doQuery(myConnection)
    myConnection.close()
except Exception as e:
    print(f"Error connecting to database: {e}")
    input("Press ENTER to exit")
